"""Excel .xlsx import/export via openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from pysheet.io.base import FormatAdapter
from pysheet.model.sheet import Sheet
from pysheet.model.workbook import Workbook


class XLSXAdapter(FormatAdapter):
    """Read and write Excel .xlsx files."""

    supported_extensions = [".xlsx", ".xlsm"]
    can_read_formulas = True
    can_write_formulas = True
    multi_sheet = True

    def read(self, path: Path, **opts: object) -> Workbook:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for .xlsx support: pip install openpyxl")

        wb_xl = openpyxl.load_workbook(str(path), data_only=False)
        wb = Workbook()
        wb.sheets.clear()
        for ws in wb_xl.worksheets:
            sheet = Sheet(name=ws.title)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    r = cell.row - 1
                    c = cell.column - 1
                    val = cell.value
                    formula: str | None = None
                    if isinstance(val, str) and val.startswith("="):
                        formula = val
                        val = None
                    sheet.set_cell_value(r, c, val, formula=formula, record_history=False)
            # Column widths
            for col_letter, col_dim in ws.column_dimensions.items():
                from openpyxl.utils import column_index_from_string
                c_idx = column_index_from_string(col_letter) - 1
                if col_dim.width:
                    sheet.set_col_width(c_idx, int(col_dim.width))
            wb.sheets.append(sheet)
        wb.filepath = path
        return wb

    def write(self, workbook: Workbook, path: Path, **opts: object) -> None:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for .xlsx support: pip install openpyxl")

        wb_xl = openpyxl.Workbook()
        wb_xl.remove(wb_xl.active)  # remove default sheet
        for sheet in workbook.sheets:
            ws = wb_xl.create_sheet(title=sheet.name)
            for (r, c), cell in sheet.cells.items():
                xl_cell = ws.cell(row=r + 1, column=c + 1)
                if cell.formula:
                    xl_cell.value = cell.formula
                else:
                    xl_cell.value = cell.value
                # Apply formatting
                from openpyxl.styles import Font, Alignment, PatternFill
                font_kwargs: dict[str, Any] = {}
                if cell.fmt.bold:
                    font_kwargs["bold"] = True
                if cell.fmt.italic:
                    font_kwargs["italic"] = True
                if cell.fmt.underline:
                    font_kwargs["underline"] = "single"
                if cell.fmt.fg_color:
                    font_kwargs["color"] = cell.fmt.fg_color.lstrip("#")
                if font_kwargs:
                    xl_cell.font = Font(**font_kwargs)
                align_map = {"left": "left", "right": "right", "center": "center"}
                xl_cell.alignment = Alignment(horizontal=align_map.get(cell.fmt.align, "right"))
                if cell.fmt.bg_color:
                    xl_cell.fill = PatternFill("solid", fgColor=cell.fmt.bg_color.lstrip("#"))
            # Column widths
            for col_idx, width in sheet.col_widths.items():
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = width
        wb_xl.save(str(path))
